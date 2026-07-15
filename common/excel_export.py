from __future__ import annotations

from io import BytesIO
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def export_queryset_to_excel(
    queryset,
    headers: list[str],
    row_builder: Callable,
    filename: str,
    sheet_title: str = "Sheet1",
) -> HttpResponse:
    """
    queryset    - eksport qilinadigan queryset (filtrlangan)
    headers     - ustun sarlavhalari ro'yxati
    row_builder - har bir obyektni qatorga aylantiruvchi funksiya: obj -> list
    filename    - yuklanadigan fayl nomi
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for obj in queryset:
        ws.append(row_builder(obj))

    # Ustun kengligini avtomatik moslashtirish
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        max_len = max(
            [len(str(header))]
            + [len(str(row[i - 1])) for row in ws.iter_rows(min_row=2, values_only=True)]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ExcelImportError(Exception):
    """Faylni o'qishda umumiy xatolik (noto'g'ri format, bo'sh fayl va h.k.)"""


def read_rows_from_excel(
    file,
    expected_headers: list[str],
) -> Iterable[dict]:
    """
    Yuklangan .xlsx faylni o'qib, har bir qatorni dict shaklida qaytaradi:
    {"header1": value1, "header2": value2, ...}

    file             - InMemoryUploadedFile (request.FILES['file'])
    expected_headers - kutilayotgan ustun sarlavhalari (birinchi qatorda bo'lishi kerak)
    """
    try:
        wb = load_workbook(filename=BytesIO(file.read()), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelImportError(f"Fayl o'qilmadi: fayl noto'g'ri formatda yoki buzilgan. ({e})")

    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise ExcelImportError("Fayl bo'sh.")

    file_headers = [str(h).strip() if h is not None else "" for h in header_row]
    missing = [h for h in expected_headers if h not in file_headers]
    if missing:
        raise ExcelImportError(
            f"Fayl sarlavhalari mos kelmadi. Yetishmayotgan ustunlar: {', '.join(missing)}"
        )

    header_index = {h: file_headers.index(h) for h in expected_headers}

    for row_number, row in enumerate(rows, start=2):
        if row is None or all(v is None for v in row):
            continue  # bo'sh qatorlarni tashlab ketamiz
        item = {h: row[idx] for h, idx in header_index.items()}
        item["_row_number"] = row_number
        yield item